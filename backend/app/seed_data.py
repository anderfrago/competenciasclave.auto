from pathlib import Path

from openpyxl import load_workbook

from .extensions import db
from .models import AppSetting, Competency, CompetencyItem, RubricLevel

COMPETENCY_LAYOUT = [
    ("Autonomía", 5),
    ("Adaptación al entorno", 6),
    ("Competencia digital", 8),
    ("Comunicación", 18),
    ("Emprendimiento / Innovación", 11),
    ("Responsabilidad", 10),
    ("Trabajo en equipo", 21),
]
REVERSE_MARKERS = ("Interrumpo", "ataco con la palabra", "Me enfado", "persona conflictiva")
DEFAULT_ENCOURAGEMENT = (
    "Gracias por dedicar este tiempo a conocerte mejor. Cada respuesta es un paso para "
    "seguir creciendo: identifica una acción pequeña y ponla en práctica esta semana."
)


def _source_path() -> Path:
    return Path(__file__).resolve().parents[2] / "Template Formulario CCs (respuestas).xlsx"


def seed_database() -> None:
    """Carga una sola vez las competencias y textos del Excel de referencia."""
    if Competency.query.count():
        if not db.session.get(AppSetting, "encouragement_message"):
            db.session.add(AppSetting(key="encouragement_message", value=DEFAULT_ENCOURAGEMENT))
            db.session.commit()
        return

    source = _source_path()
    if not source.exists():
        raise FileNotFoundError(f"No se encuentra el archivo inicial: {source.name}")

    workbook = load_workbook(source, data_only=True)
    profiles = workbook["Perfiles_CC"]
    feedback_by_competency = {}
    for row in profiles.iter_rows(min_row=2, values_only=True):
        competency, label, feedback = row[:3]
        if competency and label and feedback:
            canonical_label = "Generado" if label == "Consolidado" else label
            feedback_by_competency.setdefault(competency, {})[canonical_label] = feedback

    responses = workbook["Respuestas de formulario 1"]
    statements = [responses.cell(1, column).value for column in range(3, 82)]
    position = 0
    for order, (name, item_count) in enumerate(COMPETENCY_LAYOUT, start=1):
        competency = Competency(name=name, sort_order=order, active=True)
        db.session.add(competency)
        db.session.flush()

        for item_order, statement in enumerate(statements[position:position + item_count], start=1):
            db.session.add(
                CompetencyItem(
                    competency_id=competency.id,
                    statement=statement,
                    reverse_score=any(marker in statement for marker in REVERSE_MARKERS),
                    sort_order=item_order,
                )
            )
        position += item_count

        feedback = feedback_by_competency.get(name, {})
        for label, maximum in (("Incipiente", 2.0), ("En desarrollo", 3.0), ("Generado", 4.0)):
            db.session.add(
                RubricLevel(
                    competency_id=competency.id,
                    label=label,
                    max_score=maximum,
                    feedback=feedback.get(label, f"Resultado {label} en {name}."),
                )
            )

    db.session.add(AppSetting(key="encouragement_message", value=DEFAULT_ENCOURAGEMENT))
    db.session.commit()
